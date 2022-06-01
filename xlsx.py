import requests
import openpyxl  # excel模块，属于第三方库可能需要下载该模块，我用过CSV模块但是中文编码有些问题所以没用了，改用该模块了
import notion

token = "secret_PMHzk6jcX9zw1RMkCBjtEp45MDHxNV1JIFwMtnF0CMV"
database_id = "11bd8c08-d4cb-488c-b26f-5418ebb0f5be"
# 某个界面的id,
# 由以上数据可请求出最终界面的database_id
page_id = "b1eabf0d6fd04c949d9ed88a290a2c57"
r = requests.request(
    "GET",
    "https://api.notion.com/v1/pages/" + page_id,
    headers={"Authorization": "Bearer " + token, "Notion-Version": "2021-05-13"},
)
print(r.text)




# #  以 f开头表示在字符串内支持大括号内的python 表达式
# # path = f'{"C:/Users/valuechaintech/PycharmProjects/pythonProject/微信收藏链接.xlsx"}/tem_file/xxxx.xlsx'
# path = "C:/Users/valuechaintech/PycharmProjects/pythonProject/微信收藏链接.xlsx"
# # 加载文件对象
# wb = openpyxl.load_workbook(path)
# # 获取当前的sheet
# sh = wb.active  # 获取sheet
#
#
# # 读取指定的单元格数据
# # cell = sh.cell(row=3, column=1).value
# # print(cell)
#
# # 获取第一页
# # first_rows = sh[1]
# # for row in first_rows:
# #     print(row.value)
#
#
# # 读取表内所有数据
# # total_data = list(sh.rows)
# # for row in total_data:
# #     # print(row)
# #     for cell in row:
# #         print(cell.value)
#
#
# #获取表格所有行和列，两者都是可迭代的
# rows = sh.rows
# columns = sh.columns
# #迭代所有的行
# for row in rows:
#     line = [col.value for col in row]
#     print(line)
#     title = line[0]
#     des = line[1]
#     datetime = line[2]
#     source = line[3]
#     url = line[4]
#
#     body = {
#     "parent": {"type": "database_id", "database_id": database_id},
#     "properties": {},
#     }
#     # body111 = body_properties_input(body,'时间','rich_text',time)
#     body111 = notion.body_properties_input(body,'Name','title',title)
#     body111 = notion.body_properties_input(body,'标题','rich_text',title)
#     body111 = notion.body_properties_input(body,'描述','rich_text',des)
#     body111 = notion.body_properties_input(body,'时间','rich_text',datetime)
#     body111 = notion.body_properties_input(body,'来源','rich_text',source)
#     body111 = notion.body_properties_input(body,'url','rich_text',url)
#     body.update(body111)
#
#     url_notion_additem = 'https://api.notion.com/v1/pages'
#     notion_additem = requests.post(url_notion_additem,  headers={"Authorization": "Bearer " + token, "Notion-Version": "2021-05-13"}, json=body)
#
#     print(notion_additem.text)
#
#     if notion_additem.status_code == 200:
#         print('·更新成功')
#     else:
#         print('·更新失败')
#
#


# cell = sheet.cell(row=1, column=1).value

# # 读取xlsx文件
# workbook = load_workbook(path)
# # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
# sheetnames = workbook.sheetnames
# print(sheetnames)
# # 默认只能获取第一个工作表
# table = workbook.active
# for sheetname in sheetnames:
#     table = workbook[sheetname]
#     # table = workbook.get_sheet_by_name(sheetname)
#     rows = table.max_row
#     cols = table.max_column
#
#     for row in range(rows):
#         for col in range(cols):
#             data = table.cell(row + 1, col + 1).value
#             print(data)
